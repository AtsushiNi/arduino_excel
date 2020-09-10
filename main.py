import serial # シリアル通信用ライブラリ
import re # データを正規表現で切り分けるのに使うライブラリ
import numpy as np # 配列の操作用ライブラリ
import openpyxl # Excel操作用のライブラリ
import atexit # プログラム終了時に関数を呼び出す仕組み

# 定数
PORT = 'COM5' # ArduinoのCOMポート
LOG_FILE_NAME = 'logs/log.xlsx' # ログを保存するファイル

# データを格納する配列
number = np.empty(0)
time = np.empty(0)
value1 = np.empty(0)
value2 = np.empty(0)

# Ctrl+Cで終了時にExcelにデータを記録する
def saveData():
  try:
    print("データを保存します。しばらくお待ちください...")
    # Excelの準備
    wb = openpyxl.load_workbook(LOG_FILE_NAME)
    wb.remove(wb["Sheet1"])
    sheet = wb.create_sheet("Sheet1")

    # 1行目
    sheet.cell(row=1,column=1,value="number")
    sheet.cell(row=1,column=2,value="time")
    sheet.cell(row=1,column=3,value="value1")
    sheet.cell(row=1,column=4,value="value2")

    for i in range(np.size(number)-1):
      sheet.cell(row=i+2,column=1,value=number[i])
      sheet.cell(row=i+2,column=2,value=time[i])
      sheet.cell(row=i+2,column=3,value=value1[i])
      sheet.cell(row=i+2,column=4,value=value2[i])

    # 保存
    wb.save(LOG_FILE_NAME)

    print("データ保存完了")  
  except Exception as e:
    print("ERROR!!!")
    print(e)

atexit.register(saveData)

ser = serial.Serial(PORT, 9600) # ボーレートは9600にしている

while True:
  line = ser.readline()
  try:
    # ここからは状況次第で書き換えて!
    string_line = line.decode().rstrip('\r\n') # バイト文字から文字列へ変換した後、改行コードを削除
    data_sets = re.split(',', string_line) # data_sets: 1行分のデータの配列

    # プログラム開始時や通信不良で、データセットの途中から処理すると配列の順番がズレるので、スキップする
    if re.split(':', data_sets[0])[0] != 'number':
      continue
    if np.size(data_sets) != 4:
      continue

    print(data_sets)

    for data_set in data_sets:
      data = re.split(':', data_set) # data[0]: 値の名前, data[1]: 値

      if data[0] == 'number':
        number = np.append(number, int(data[1]))
      elif data[0] == 'time':
        time = np.append(time, int(data[1]))
      elif data[0] == 'value1':
        value1 = np.append(value1, float(data[1]))
      elif data[0] == 'value2':
        value2 = np.append(value2, float(data[1]))

  except UnicodeDecodeError: # 受信したデータがおかしいときのエラー処理
    print('ERROR!!!')
    print(line)